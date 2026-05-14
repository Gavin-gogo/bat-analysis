import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="蝙蝠偵測分析工具",
    page_icon="🦇",
    layout="wide",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;500;700&family=Space+Mono:wght@400;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Noto Sans TC', sans-serif;
}

.stApp {
    background: #0f1117;
    color: #e8eaf0;
}

/* Header */
.hero {
    background: linear-gradient(135deg, #1a1f35 0%, #0d1b2a 50%, #1a1f35 100%);
    border: 1px solid #2a3550;
    border-radius: 16px;
    padding: 2.5rem 3rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero::before {
    content: '';
    position: absolute;
    top: -50%; left: -50%;
    width: 200%; height: 200%;
    background: radial-gradient(circle at 70% 30%, rgba(99,179,237,0.06) 0%, transparent 60%);
    pointer-events: none;
}
.hero h1 {
    font-family: 'Space Mono', monospace;
    font-size: 2rem;
    color: #63b3ed;
    margin: 0 0 0.5rem 0;
    letter-spacing: -0.5px;
}
.hero p {
    color: #8896b3;
    font-size: 0.95rem;
    margin: 0;
}

/* Upload zone */
.upload-zone {
    background: #161b2e;
    border: 2px dashed #2a3a5e;
    border-radius: 12px;
    padding: 2rem;
    text-align: center;
    transition: border-color 0.2s;
}
.upload-zone:hover { border-color: #63b3ed; }

/* Metric cards */
.metric-row { display: flex; gap: 1rem; margin: 1rem 0; }
.metric-card {
    flex: 1;
    background: #161b2e;
    border: 1px solid #2a3550;
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
}
.metric-card .label {
    font-size: 0.78rem;
    color: #6b7a99;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-bottom: 0.4rem;
}
.metric-card .value {
    font-family: 'Space Mono', monospace;
    font-size: 2rem;
    font-weight: 700;
    color: #63b3ed;
}
.metric-card .sub { font-size: 0.8rem; color: #8896b3; }

/* Section headers */
.section-header {
    font-family: 'Space Mono', monospace;
    font-size: 0.85rem;
    color: #63b3ed;
    text-transform: uppercase;
    letter-spacing: 2px;
    border-bottom: 1px solid #2a3550;
    padding-bottom: 0.6rem;
    margin: 2rem 0 1rem 0;
}

/* Dataframe styling */
.stDataFrame { border-radius: 10px; overflow: hidden; }

/* Download button */
.stDownloadButton > button {
    background: linear-gradient(135deg, #2b6cb0, #2c5282) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.6rem 2rem !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.85rem !important;
    letter-spacing: 0.5px !important;
    transition: opacity 0.2s !important;
}
.stDownloadButton > button:hover { opacity: 0.85 !important; }

/* Badge */
.badge {
    display: inline-block;
    background: #1e3a5f;
    color: #63b3ed;
    border-radius: 999px;
    padding: 0.15rem 0.7rem;
    font-size: 0.75rem;
    font-family: 'Space Mono', monospace;
    margin-left: 0.5rem;
}

.tag-true { color: #68d391; font-weight: 700; }
.tag-false { color: #718096; }

/* Sidebar */
section[data-testid="stSidebar"] {
    background: #0d1117 !important;
    border-right: 1px solid #1e2433;
}
</style>
""", unsafe_allow_html=True)


# ── Core analysis functions ────────────────────────────────────────────────────

def compute_concurrent(df_all: pd.DataFrame, source_zx: pd.DataFrame, source_cx: pd.DataFrame,
                        threshold: float = 2.0, add_pair_id: bool = False) -> pd.DataFrame:
    """Mark rows where 正下 and 側向 video_time_sec differ < threshold seconds.
    If add_pair_id=True, also adds a '配對組' column (int) so each matched pair shares the same group number.
    """
    times_zx = source_zx['video_time_sec'].values
    times_cx = source_cx['video_time_sec'].values

    # Build list of all pairs
    pairs = []  # [(zx_pos, cx_pos), ...]
    for i, ta in enumerate(times_zx):
        for j, tb in enumerate(times_cx):
            if abs(float(ta) - float(tb)) < threshold:
                pairs.append((i, j))

    c_zx = {p[0] for p in pairs}
    c_cx = {p[1] for p in pairs}

    df_all = df_all.copy()
    df_all['同時出現'] = ''
    zx_idx = df_all[df_all['角度'] == '正下'].index.tolist()
    cx_idx = df_all[df_all['角度'] == '側向'].index.tolist()

    for pos, idx in enumerate(zx_idx):
        if pos in c_zx:
            df_all.at[idx, '同時出現'] = 'true'
    for pos, idx in enumerate(cx_idx):
        if pos in c_cx:
            df_all.at[idx, '同時出現'] = 'true'

    if add_pair_id:
        # Assign a group number to each unique pair; rows in the same pair share the same group id.
        # A row may belong to multiple pairs → assign the first (lowest) group id.
        df_all['配對組'] = pd.NA

        # Map: zx_pos → list of group ids,  cx_pos → list of group ids
        zx_to_group = {}
        cx_to_group = {}
        for group_id, (zx_pos, cx_pos) in enumerate(pairs):
            zx_to_group.setdefault(zx_pos, group_id)
            cx_to_group.setdefault(cx_pos, group_id)

        for pos, idx in enumerate(zx_idx):
            if pos in zx_to_group:
                df_all.at[idx, '配對組'] = zx_to_group[pos]
        for pos, idx in enumerate(cx_idx):
            if pos in cx_to_group:
                df_all.at[idx, '配對組'] = cx_to_group[pos]

    return df_all, pairs


def run_analysis(uploaded_file, sheet_zx: str, sheet_cx: str, concurrent_threshold: float):
    """Main analysis pipeline. Returns (df1, df2, pivot, bat_conc, bird_conc)."""
    xl = pd.read_excel(uploaded_file, sheet_name=None)

    if sheet_zx not in xl or sheet_cx not in xl:
        return None, None, None, None, None, f"找不到工作表：請確認工作表名稱（{sheet_zx} / {sheet_cx}）"

    df_zx = xl[sheet_zx].copy()
    df_cx = xl[sheet_cx].copy()

    required_cols = {'video_time_sec', 'CP', '角度', '物種', '高度', 'bat_count'}
    for label, df in [(sheet_zx, df_zx), (sheet_cx, df_cx)]:
        missing = required_cols - set(df.columns)
        if missing:
            return None, None, None, None, None, f"工作表「{label}」缺少欄位：{missing}"

    # Sheet 1: all data
    df1 = pd.concat([df_zx, df_cx], ignore_index=True).sort_values('video_time_sec').reset_index(drop=True)
    df1, _ = compute_concurrent(df1, df_zx.reset_index(drop=True), df_cx.reset_index(drop=True), concurrent_threshold)

    # Sheet 2: CP=0
    df2_zx = df_zx[df_zx['CP'] == 0].reset_index(drop=True)
    df2_cx = df_cx[df_cx['CP'] == 0].reset_index(drop=True)
    df2 = pd.concat([df2_zx, df2_cx], ignore_index=True).sort_values('video_time_sec').reset_index(drop=True)
    df2, pairs2 = compute_concurrent(df2, df2_zx, df2_cx, concurrent_threshold, add_pair_id=True)

    # 新增「人工驗證」欄位（下拉選單 true/false，預設空白），緊接在「同時出現」後面
    # 只有同時出現=true 的列才需要驗證
    conc_col_pos = df2.columns.tolist().index('同時出現')
    df2.insert(conc_col_pos + 1, '人工驗證', '')

    # Sheet 3 pivot
    pivot = (
        df2.groupby(['角度', '物種', '高度'])['bat_count']
        .sum()
        .reset_index()
        .rename(columns={'bat_count': '數量(bat_count加總)'})
        .sort_values(['角度', '物種', '高度'])
    )

    # 同時出現：每組配對取兩筆 bat_count 的最大值，再依物種加總
    zx_rows = df2_zx.reset_index(drop=True)
    cx_rows = df2_cx.reset_index(drop=True)

    conc_pairs = {}  # (i,j) -> (sp_zx, sp_cx, max_bat_count)
    for i, row_zx in zx_rows.iterrows():
        for j, row_cx in cx_rows.iterrows():
            if abs(float(row_zx['video_time_sec']) - float(row_cx['video_time_sec'])) < concurrent_threshold:
                pair_key = (i, j)
                if pair_key not in conc_pairs:
                    conc_pairs[pair_key] = (
                        row_zx['物種'],
                        row_cx['物種'],
                        max(int(row_zx['bat_count']), int(row_cx['bat_count']))
                    )

    # 依物種加總各配對的最大值
    species_total = {}
    for (sp_zx, sp_cx, pair_max) in conc_pairs.values():
        if sp_zx == sp_cx:
            species_total[sp_zx] = species_total.get(sp_zx, 0) + pair_max
        else:
            species_total[sp_zx] = species_total.get(sp_zx, 0) + pair_max
            species_total[sp_cx] = species_total.get(sp_cx, 0) + pair_max

    bat_conc = int(species_total.get('蝙蝠', 0))
    bird_conc = int(species_total.get('鳥', 0))

    # 所有物種的加總（含蝙蝠、鳥及其他）
    conc_species_total = species_total

    return df1, df2, pairs2, pivot, bat_conc, bird_conc, conc_species_total, None


def build_excel(df1: pd.DataFrame, df2: pd.DataFrame, pairs2: list,
                pivot: pd.DataFrame, bat_conc: int, bird_conc: int,
                conc_species_total: dict) -> bytes:
    """Build the 3-sheet xlsx and return as bytes."""
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    H_FILL = PatternFill('solid', start_color='2E75B6')
    H_FILL2 = PatternFill('solid', start_color='70AD47')
    D_FONT = Font(name='Arial', size=10)

    # Pair row colors: alternate between white and light green
    PAIR_FILLS = [
        PatternFill('solid', start_color='FFFFFF'),   # white
        PatternFill('solid', start_color='D9F0D3'),   # light green
    ]

    def auto_width(ws):
        for col in ws.columns:
            mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 42)

    def write_df(ws, df):
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(1, ci, col)
            c.font = H_FONT; c.fill = H_FILL
            c.alignment = Alignment(horizontal='center')
            c.border = thin
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val if pd.notna(val) else '')
                c.font = D_FONT; c.border = thin
        auto_width(ws)

    # Sheet 1
    ws1 = wb.active
    ws1.title = '全部資料合併'
    write_df(ws1, df1)

    # ── Sheet 2: CP=0，含人工驗證欄、配對底色 ──────────────────────────────
    ws2 = wb.create_sheet('CP=0資料合併')

    # Build row→group_id mapping from df2's '配對組' column
    # df2 has '配對組' as a column (int or pd.NA); we use it for coloring
    group_col = '配對組'
    has_group = group_col in df2.columns

    # Determine which df2 column index is '人工驗證'
    cols = df2.columns.tolist()
    # Drop '配對組' from output (internal use only)
    output_cols = [c for c in cols if c != group_col]
    manval_col_idx = output_cols.index('人工驗證') + 1  # 1-based Excel col

    # Write header
    for ci, col in enumerate(output_cols, 1):
        c = ws2.cell(1, ci, col)
        c.font = H_FONT; c.fill = H_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    # Assign alternating color per group_id
    # group_id → color index (0 or 1)
    group_color_map = {}
    color_counter = 0

    for ri, row_data in enumerate(df2.itertuples(index=False), 2):
        row_dict = dict(zip(df2.columns, row_data))
        gid = row_dict.get(group_col, pd.NA)

        # Determine fill color
        if pd.notna(gid):
            gid_int = int(gid)
            if gid_int not in group_color_map:
                group_color_map[gid_int] = color_counter % 2
                color_counter += 1
            fill = PAIR_FILLS[group_color_map[gid_int]]
        else:
            fill = None  # no color for non-concurrent rows

        for ci, col in enumerate(output_cols, 1):
            val = row_dict[col]
            c = ws2.cell(ri, ci, val if pd.notna(val) else '')
            c.font = D_FONT; c.border = thin
            if fill:
                c.fill = fill

    # Add data validation (dropdown: true / false) on '人工驗證' column
    dv = DataValidation(
        type="list",
        formula1='"true,false"',
        allow_blank=True,
        showDropDown=False,  # False = show the dropdown arrow
        showErrorMessage=True,
        errorTitle='輸入錯誤',
        error='請選擇 true 或 false',
    )
    last_row = len(df2) + 1
    dv.sqref = f"{get_column_letter(manval_col_idx)}2:{get_column_letter(manval_col_idx)}{last_row}"
    ws2.add_data_validation(dv)

    auto_width(ws2)

    # Sheet 3
    ws3 = wb.create_sheet('總成果表')
    ws3.merge_cells('A1:D1')
    ws3['A1'] = '總成果表（CP=0）'
    ws3['A1'].font = Font(bold=True, name='Arial', size=14, color='1F3864')
    ws3['A1'].alignment = Alignment(horizontal='center')

    row = 3
    ws3.cell(row, 1, '各角度 / 物種 / 高度 數量統計（bat_count 加總）').font = Font(bold=True, name='Arial', size=11)
    row += 1
    for ci, h in enumerate(['角度', '物種', '高度', '數量(bat_count加總)'], 1):
        c = ws3.cell(row, ci, h)
        c.font = H_FONT; c.fill = H_FILL
        c.alignment = Alignment(horizontal='center'); c.border = thin
    row += 1
    for _, r in pivot.iterrows():
        vals = [r['角度'], r['物種'], int(r['高度']), int(r['數量(bat_count加總)'])]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row, ci, val)
            c.font = D_FONT; c.border = thin; c.alignment = Alignment(horizontal='center')
        row += 1
    total = int(pivot['數量(bat_count加總)'].sum())
    for ci, val in enumerate(['合計', '', '', total], 1):
        c = ws3.cell(row, ci, val)
        c.font = Font(bold=True, name='Arial'); c.border = thin
        c.fill = PatternFill('solid', start_color='BDD7EE')
        c.alignment = Alignment(horizontal='center')
    row += 2

    ws3.cell(row, 1, '同時出現（正下 ↔ 側向 video_time_sec 相差 < 2 秒）各配對取 bat_count 最大值後加總').font = Font(bold=True, name='Arial', size=11)
    row += 1
    for ci, h in enumerate(['物種', '數量（配對最大值加總）'], 1):
        c = ws3.cell(row, ci, h)
        c.font = H_FONT; c.fill = H_FILL2
        c.alignment = Alignment(horizontal='center'); c.border = thin
    row += 1
    grand_total = 0
    for sp, cnt in conc_species_total.items():
        ws3.cell(row, 1, sp).font = D_FONT; ws3.cell(row, 1).border = thin; ws3.cell(row, 1).alignment = Alignment(horizontal='center')
        ws3.cell(row, 2, cnt).font = D_FONT; ws3.cell(row, 2).border = thin; ws3.cell(row, 2).alignment = Alignment(horizontal='center')
        grand_total += cnt
        row += 1
    # Grand total row
    for ci, val in enumerate(['合計', grand_total], 1):
        c = ws3.cell(row, ci, val)
        c.font = Font(bold=True, name='Arial'); c.border = thin
        c.fill = PatternFill('solid', start_color='C6EFCE')
        c.alignment = Alignment(horizontal='center')
    row += 1
    auto_width(ws3)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="hero">
    <h1>🦇 蝙蝠偵測分析工具</h1>
    <p>上傳熱像儀偵測結果 Excel，自動合併正下／側向資料、標記同時出現事件、輸出三頁分析報表。</p>
</div>
""", unsafe_allow_html=True)

# Sidebar settings
with st.sidebar:
    st.markdown("### ⚙️ 分析設定")
    st.markdown("---")
    sheet_zx = st.text_input("正下 工作表名稱", value="正下")
    sheet_cx = st.text_input("側向 工作表名稱", value="側向")
    threshold = st.slider(
        "同時出現閾值（秒）",
        min_value=0.5, max_value=10.0, value=2.0, step=0.5,
        help="正下與側向 video_time_sec 相差小於此值，即標記為同時出現"
    )
    st.markdown("---")
    st.markdown("**所需欄位**")
    st.markdown("""
    - `角度`
    - `video_time_sec`
    - `物種`
    - `bat_count`
    - `CP`
    - `高度`
    """)

# File upload
st.markdown('<div class="section-header">上傳檔案</div>', unsafe_allow_html=True)
uploaded = st.file_uploader("選擇 Excel 檔案（.xlsx）", type=["xlsx"], label_visibility="collapsed")

if uploaded:
    with st.spinner("分析中..."):
        df1, df2, pairs2, pivot, bat_conc, bird_conc, conc_species_total, err = run_analysis(uploaded, sheet_zx, sheet_cx, threshold)

    if err:
        st.error(f"❌ {err}")
    else:
        # ── Metrics ──────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">分析結果概覽</div>', unsafe_allow_html=True)
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.metric("全部資料（筆）", f"{len(df1):,}")
        with c2:
            st.metric("CP=0 資料（筆）", f"{len(df2):,}")
        with c3:
            st.metric("同時出現（筆）", f"{(df2['同時出現']=='true').sum()}")
        with c4:
            st.metric("同時出現 蝙蝠", f"{bat_conc} 隻")
        with c5:
            st.metric("同時出現 鳥", f"{bird_conc} 隻")

        # ── Tabs ─────────────────────────────────────────────────────────────
        tab1, tab2, tab3 = st.tabs(["📋 工作表1：全部資料", "📋 工作表2：CP=0", "📊 工作表3：總成果表"])

        with tab1:
            st.markdown(f"共 **{len(df1):,}** 筆，同時出現標記 **{(df1['同時出現']=='true').sum()}** 筆")
            st.dataframe(df1, use_container_width=True, height=400)

        with tab2:
            st.markdown(f"共 **{len(df2):,}** 筆，bat_count 總和 **{int(df2['bat_count'].sum())}**，同時出現標記 **{(df2['同時出現']=='true').sum()}** 筆")
            display_df2 = df2.drop(columns=['配對組'], errors='ignore')
            st.dataframe(display_df2, use_container_width=True, height=400)

        with tab3:
            col_a, col_b = st.columns([2, 1])
            with col_a:
                st.markdown("#### 各角度／物種／高度數量（bat_count加總）")
                st.dataframe(pivot, use_container_width=True)
                st.markdown(f"**合計：{int(pivot['數量(bat_count加總)'].sum())} 隻**")
            with col_b:
                st.markdown("#### 同時出現（配對最大值加總）")
                conc_df = pd.DataFrame({
                    '物種': list(conc_species_total.keys()) + ['合計'],
                    '數量': list(conc_species_total.values()) + [sum(conc_species_total.values())]
                })
                st.dataframe(conc_df, use_container_width=True, hide_index=True)

        # ── Download ─────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">下載結果</div>', unsafe_allow_html=True)
        xlsx_bytes = build_excel(df1, df2, pairs2, pivot, bat_conc, bird_conc, conc_species_total)
        original_name = uploaded.name.replace('.xlsx', '')
        st.download_button(
            label="⬇️  下載分析結果 Excel",
            data=xlsx_bytes,
            file_name=f"{original_name}_分析結果.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:
    st.info("👆 請在上方上傳 Excel 檔案以開始分析")
    st.markdown("""
    **使用流程：**
    1. 在左側設定工作表名稱（預設：正下 / 側向）與同時出現閾值
    2. 上傳 .xlsx 檔案
    3. 預覽三個工作表的分析結果
    4. 下載輸出的 Excel 報表
    """)
